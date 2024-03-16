# Module Untuk Membaca excel
import openpyxl
from openpyxl import Workbook
read = openpyxl.load_workbook('datapml.xlsx')
write = Workbook()
sheet_read = read.active
sheet_write = write.active

# Inisasi list menampung data, jumlah row dan column
Data = []
data_row = sheet_read.max_row
data_column = sheet_read.max_column

# Membaca data dan menyimpan ke list
for a in range (1, sheet_read.max_row + 1):
    dummylist = []
    for b in range (1, sheet_read.max_column + 1):
        dummylist.append(float(sheet_read.cell(row=a, column=b).value))
    Data.append(dummylist)
print(Data)

for k in range (2,3):

    # Penghitungan Kmeans dengan data yang dimasukkan manual ke dalam codingan
    # Data dengan dimensi X,Y serta 2 cluster (bisa ditambah dengan sedikit modifikasi)

    import random
    import math

    Perulangan = 0

    # Data Bisa Diubah
    # Data = [[2, 0], [1, 3], [3, 5], [2, 2], [4, 6]]
    # Data = [[1, 1], [1.5, 2], [3, 4], [5, 7], [3.5, 5], [4.5, 5], [3.5, 4.5]]

    # Centroid = [[1, 3], [2, 2]]
    Centroid = []

    # print(Data[0].index(min(Data[0])))    Cari index data terkecil dari array 2D

    # Pilih Centroid secara Random (jumlah centroid/cluster bisa diubah untuk skrng pakai 2 cluster)
    while len(Centroid) != 2:
        dum = Data[random.randint(0, len(Data) - 1)]
        if dum not in Centroid:
            Centroid.append(dum)
        elif dum in Centroid:
            continue

    print("Centroid Awal :", Centroid)

    # Mulai penghitungan Kmeans
    while True:
        Perulangan = Perulangan + 1

        # A adalah Cluster 1
        A = []

        # B adalah Cluster 2
        B = []

        ScrapEcludian = []

        # Hitung Ecludian Distance untuk pengelompokkan ke cluster terdekat yang mana
        # (Jika jumlah cluster diubah maka disini perlu ditambah)
        for d in Data:
            centroidA = math.sqrt((math.pow(d[0] - Centroid[0][0], 2)) + (math.pow(d[1] - Centroid[0][1], 2)))
            centroidB = math.sqrt((math.pow(d[0] - Centroid[1][0], 2)) + (math.pow(d[1] - Centroid[1][1], 2)))
            ScrapEcludian.append([centroidA, centroidB])
            if (centroidA < centroidB):
                A.append(d)
            else:
                B.append(d)
        print("Perulangan ke :", Perulangan)
        print("Team A :", A)
        print("Team B :", B)
        print("Scrap :", ScrapEcludian)
        sheet_write.insert_rows(10,1)
        for i in ScrapEcludian:
            sheet_write.append(i)
        for i in A:
            sheet_write.append(i)
        for i in B:
            sheet_write.append(i)

        # Hitung Means
        # (Jika jumlah cluster diubah maka disini perlu ditambah)
        XTotalA = 0
        YTotalA = 0
        XTotalB = 0
        YTotalB = 0
        for i in A:
            XTotalA = XTotalA + i[0]
            YTotalA = YTotalA + i[1]
        for i in B:
            XTotalB = XTotalB + i[0]
            YTotalB = YTotalB + i[1]

        XMeanA = XTotalA / len(A)
        YMeanA = YTotalA / len(A)

        XMeanB = XTotalB / len(B)
        YMeanB = YTotalB / len(B)

        DumCentroid = []
        DumCentroid.append([XMeanA, YMeanA])
        DumCentroid.append([XMeanB, YMeanB])

        if ((Centroid[0] == DumCentroid[0] or Centroid[1] == DumCentroid[0]) and (
                Centroid[0] == DumCentroid[1] or Centroid[1] == DumCentroid[1])):
            print("Final :", Centroid)
            break

        else:
            Centroid = DumCentroid
            print("Centroid :", Centroid, "\n")
write.save('test.xlsx')